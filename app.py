######################################################################################################
                                 # importar bibliotecas
######################################################################################################

import streamlit as st
from streamlit import caching
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import pandas as pd
import numpy as np
#import json
#import smtplib
from datetime import  datetime, time, timedelta
import pytz
import base64
from io import StringIO, BytesIO
# import pymongo
# from st_aggrid import AgGrid
# from pylogix import PLC
from PIL import Image
import io
import matplotlib.pyplot as plt
import cv2
from pyzbar.pyzbar import decode
import time
import qrcode
from PIL import Image
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as Image_openpyxl
from openpyxl.styles import Font, Color

#from webcam import webcam
import asyncio
import logging
import queue
import threading
import urllib.request
from pathlib import Path
from typing import List, NamedTuple

try:
    from typing import Literal
except ImportError:
    from typing_extensions import Literal  # type: ignore

import av
import cv2
import matplotlib.pyplot as plt
import numpy as np
#import pydub
import streamlit as st
from aiortc.contrib.media import MediaPlayer

from streamlit_webrtc import (
    AudioProcessorBase,
    RTCConfiguration,
    VideoProcessorBase,
    WebRtcMode,
    webrtc_streamer,
)

RTC_CONFIGURATION = RTCConfiguration(
    {"iceServers": [{"urls": ["stun:stun.l.google.com:19302"]}]}
)


from google.cloud import firestore
from google.oauth2 import service_account

key_dict = json.loads(st.secrets['textkey'])
creds = service_account.Credentials.from_service_account_info(key_dict)
db = firestore.Client(credentials=creds, project='logistica-invent')


######################################################################################################
                               #Funções
######################################################################################################

st.set_page_config(
     page_title="Inventário Logístico",
)

m = st.markdown("""
<style>
div.stButton > button:first-child{
    width: 100%;
    font-size: 18px;
}
label.css-qrbaxs{
    font-size: 18px;
}
p{
    font-size: 18px;
}
h1{
    text-align: center;
}
div.block-container{
    padding-top: 1rem;
}
div.streamlit-expanderHeader{
    width: 100%;
    font-size: 18px;
    background-color: rgb(240,242,246);
    color: black;
}
</style>""", unsafe_allow_html=True) #    font-weight: bold;


def read_barcodes(frame):

    barcodes = decode(frame)
    for barcode in barcodes:
        x, y , w, h = barcode.rect        #1
        barcode_info = barcode.data.decode('utf-8')             
        return barcode_info


def entrada_bobinas() -> None:
    st.subheader('Inserir bobina')

    dict_data = {}

    with st.form(key='myform', clear_on_submit=True):
        texto_qrcode = ''

        dict_descricao_bobinas = {
        'BOBINA ALUMINIO LATA 16 OZ COIL 00098': 50761710,
        'BOBINA ALUMINIO LATA 12 OZ COIL 00098': 50679811,
        'BOBINA ALUMINIO LATA 12 OZ COIL 98 SCRAP': 40011008,
        'BOBINA ALUMINIO LACRE PRETO': 50552903,
        'BOBINA ALUMINIO LACRE AZUL': 50527602,
        'BOBINA ALUMINIO LATA 16 OZ': 50490762,
        'BOBINA ALUMINIO LATA 12 OZ': 50490761,
        'BOBINA ALUMINIO TAMPA PRATA REFRIG.': 50490760,
        'BOBINA ALUMINIO TAMPA DOURADO CERVEJA': 50490599,
        'BOBINA ALUMINIO LACRE PRATA': 50490598,
        'BOBINA ALUMINIO LATA 12 OZ SCRAP': 40010824,
        'BOBINA ALUMINIO TAMPA BRANCA': 50527252,
        'BOBINA ALUMINIO LACRE DOURADO': 50771048,
        'BOBINA ALUMÍNIO LATA 12 OZ SCRAP': 40010824}

        tipo_bobinas = ['L3', 'L2', 'L1', 'M', 'H1', 'H2', 'H3']

        dict_data['status'] = st.selectbox('Status da bobina', ['Liberado', 'Não conforme']) # data
        dict_data['descricao'] = st.selectbox('Descrição:', list(dict_descricao_bobinas.keys()))
        dict_data['conferente'] = st.text_input('Conferente:')
        dict_data['quantidade'] = st.number_input('Quantidade:', format='%i', step=1, value=9000)
        dict_data['lote'] = st.text_input('Lote:')
        dict_data['tipo'] = st.selectbox('Tipo', tipo_bobinas)
        dict_data['data'] = st.date_input('Data entrada:')

        submit_button = st.form_submit_button(label='Salvar bobina')

        if submit_button:

            if (dict_data['conferente'] == '') or (dict_data['lote'] == ''):
                st.error('Preencha todos os campos')
            else:
                dict_data['descricao'] = dict_data['descricao'].replace(',',' ') 
                dict_data['conferente'] = dict_data['conferente'].replace(',',' ')  
                dict_data['lote'] = dict_data['lote'].replace(',',' ')  
    
                if dict_data['status'] == 'Não conforme':
                    dict_data['tipo_de_etiqueta'] = 'BLOQUEADO'

                if dict_data['status'] == 'Liberado':
                    dict_data['tipo_de_etiqueta'] = 'LIBERADO'

                dict_data['sap'] = dict_descricao_bobinas[dict_data['descricao']]

                doc_ref = db.collection('bobinas').document('bobinas')
                doc = doc_ref.get()

                if doc.exists:
                    dicionario = doc.to_dict()
                    csv = dicionario['dataframe']

                    csv_string = StringIO(csv)
                    df_bobinas = pd.read_csv(csv_string, sep=',')

                    df_bobinas = df_bobinas.append(dict_data, ignore_index=True)
                    df_bobinas.drop_duplicates(inplace=True)

                    dados = {}
                    dados['dataframe'] = df_bobinas.to_csv(index=False)

                    try:
                        doc_ref.set(dados)
                        st.success('Bobina inserida com sucesso')
                    except:
                        st.error('Erro ao inserir bobina')
 
                else:
                    df_bobinas = pd.DataFrame(dict_data, index=[0])
                    df_bobinas.drop_duplicates(inplace=True)

                    dados = {}
                    dados['dataframe'] = df_bobinas.to_csv(index=False)

                    try:
                        doc_ref.set(dados)
                        st.success('Bobina inserida com sucesso')
                    except:
                        st.error('Erro ao inserir bobina')

                time.sleep(2)
                st.experimental_rerun()


@st.cache(allow_output_mutation=True)
def read_cv2():
    return cv2.VideoCapture(0)


def visualizar_inventario() -> None:
    st.subheader('Inventários realizados')
    doc_ref = db.collection('inventario').document('inventario')
    doc = doc_ref.get()

    numero_dias = st.number_input('Quantidade de dias para busca:', value=7, step=1, format='%i')

    if doc.exists:
        dicionario = doc.to_dict()
        csv = dicionario['dataframe']

        csv_string = StringIO(csv)
        df_bobinas = pd.read_csv(csv_string, sep=',') 

        df_bobinas['status'] = df_bobinas['status'].apply(lambda x: 'Não conforme' if x != 'Liberado' else 'Liberado')
        df_bobinas['id'] = df_bobinas['nome_inventario'].astype(str) + '_' + df_bobinas['data_inventario'].astype(str)
        df_bobinas['data_inventario'] = pd.to_datetime(df_bobinas['data_inventario'], format='%d/%m/%Y').dt.date
        df_bobinas['data'] = pd.to_datetime(df_bobinas['data']).dt.date
        df_bobinas.sort_values(by=['data_inventario'], ascending=False, inplace=True)

        df_bobinas = df_bobinas.loc[(df_bobinas['data_inventario'] >= datetime.now().date() - timedelta(days=numero_dias))
                                    & (df_bobinas['data_inventario'] <= datetime.now().date())]

        lista_inventarios = df_bobinas['id'].unique()

        for inventario in lista_inventarios:
            df_inventario = df_bobinas[df_bobinas['id'] == inventario]

            df_inventario_att = df_inventario.drop_duplicates()
            df_inventario_att = df_inventario_att.sort_values(by=['data'], ascending=True)
            df_inventario_att = df_inventario_att.reset_index(drop=True)

            # st.write(df_inventario_att['data'].dtypes)
            df_inventario_att['data'] = pd.to_datetime(df_inventario_att['data']).dt.strftime('%d/%m/%Y')   #.astype(str) #strftime('%d/%m/%Y')
            df_inventario_att['data_inventario'] = pd.to_datetime(df_inventario_att['data_inventario']).dt.strftime('%d/%m/%Y')   #.astype(str) #strftime('%d/%m/%Y')
            #df_inventario_att['data'] = str(df_inventario_att['data'].split('-')[2]) + '/' + str(df_inventario_att['data'].split('-')[1]) + '/' + str(df_inventario_att['data'].split('-')[0]) #strftime('%d/%m/%Y')
            #df_inventario_att['data_inventario'] = df_inventario_att['data_inventario'].dt.strftime('%d/%m/%Y')
            

            with st.expander(f'{inventario} ({str(df_inventario_att.shape[0])})'):
                st.dataframe(df_inventario_att)
                df_xlsx = download_inventario(df_inventario_att)
                st.download_button(label = f'Download inventário {inventario}',
                                    data = df_xlsx,
                                    file_name = 'Contagem_bobinas.xlsx')
    else:
        st.warning('Não foram realizados inventários')


def download_inventario(df_inventario: pd.DataFrame) -> None:
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df_inventario.to_excel(writer, index=False, sheet_name='Inventário Bobinas')
    workbook = writer.book
    worksheet = writer.sheets['Inventário Bobinas']
    writer.save()
    processed_data = output.getvalue()
    return processed_data


def VideoProcessor(dataframe_string: str) -> None:
    class video_processor(VideoProcessorBase):

        def __init__(self):
            self.result_queue = queue.Queue()
        
        def recv(self, frame):
            img = frame.to_ndarray(format='bgr24') #bgr24
            data = ''

            barcodes = decode(img)
            for barcode in barcodes:
                x, y , w, h = barcode.rect        #1
                cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 5)
                data = barcode.data.decode('utf-8')

            if data != '' and data is not None:
                self.result_queue.put(data)               

            return av.VideoFrame.from_ndarray(img, format='bgr24')

    webrtc_ctx = webrtc_streamer(key='opencv-filter',
        video_processor_factory=video_processor,
        mode=WebRtcMode.SENDRECV,
        rtc_configuration=RTC_CONFIGURATION,
        media_stream_constraints={"video": True, "audio": False},
        async_processing=True)

    if 'data_inventario' not in st.session_state:
        st.session_state['data_inventario'] = '' 
   
    if webrtc_ctx.state.playing:
        labels_placeholder = st.empty()
        colunas = 'status,descricao,conferente,quantidade,lote,tipo,data,sap\n'

        while True:
            if webrtc_ctx.video_processor:
                try:
                    result = webrtc_ctx.video_processor.result_queue.get(timeout=2.0)
                except queue.Empty:
                    result = None
                    labels_placeholder.warning('Nenhum QR code detectado')
            else:
                break

            if result is not None:
                if (result != st.session_state.data_inventario) and (result.count(',') == 7):
                    nova_bobina_inventario = ''.join((colunas, result))
                    csv_string = StringIO(nova_bobina_inventario)
                    df_inventario_atual = pd.read_csv(csv_string, sep=',')
                    df_inventario_atual['data_inventario'] = datetime.now().strftime('%d/%m/%Y')
                    df_inventario_atual['nome_inventario'] = 'Inventario_'

                    st.write(result)
                    update_inventario(colunas, df_inventario_atual)
                    st.session_state.data_inventario = result

                if result == st.session_state.data_inventario and result.count(',') == 7:
                    labels_placeholder.info('Bobina já adicionada ao inventário')

                if result.count(',') != 7:
                    labels_placeholder.error('QR code inválido')


def inserir_inventario() -> None:
    st.subheader('Inventário de bobinas')
    VideoProcessor('colunas')


def update_inventario(colunas, df_inventario_atual):
        
        if st.session_state.data_inventario != colunas:
                doc_ref = db.collection('inventario').document('inventario')
                doc = doc_ref.get()

                if doc.exists:
                    dicionario = doc.to_dict()
                    csv = dicionario['dataframe']

                    csv_string = StringIO(csv)
                    df_bobinas = pd.read_csv(csv_string, sep=',')

                    df_bobinas = df_bobinas.append(df_inventario_atual, ignore_index=True)
                    df_bobinas.drop_duplicates(inplace=True)

                    dados = {}
                    dados['dataframe'] = df_bobinas.to_csv(index=False)

                    try:
                        doc_ref.set(dados)
                        st.session_state['data_inventario'] = colunas
                        st.success('Inventário realizado com sucesso')
                    except:
                        st.error('Erro ao salvar inventário')
                else:
                    df_bobinas = pd.DataFrame(df_inventario_atual, index=[0])
                    df_bobinas.drop_duplicates(inplace=True)

                    dados = {}
                    dados['dataframe'] = df_bobinas.to_csv(index=False)

                    try:
                        doc_ref.set(dados)
                        st.session_state['data_inventario'] = colunas
                        st.success('Inventário realizado com sucesso')
                    except:
                        st.error('Erro ao salvar inventário')

                time.sleep(1)
        else:
            st.warning('Não há bobinas para armazenar')


def download_etiqueta(texto_qrcode: str, dados_bobina: pd.DataFrame) -> None:
    # imagem_bobina_qr = qrcode.make(texto_qrcode , version=12, box_size=2, border=2, error_correction=qrcode.constants.ERROR_CORRECT_H) #, fit=True)
    # image_bytearray = io.BytesIO()
    # imagem_bobina_qr.save(image_bytearray, format='PNG', name='qrcode.png')

    if dados_bobina.loc['tipo_de_etiqueta'] == 'LIBERADO':
        imagem_bobina_qr = qrcode.make(texto_qrcode , version=15, box_size=2, border=2, error_correction=qrcode.constants.ERROR_CORRECT_H) #, fit=True)
        image_bytearray = io.BytesIO()
        imagem_bobina_qr.save(image_bytearray, format='PNG', name='qrcode.png')
        wb = load_workbook('LIBERADO.xlsx')
        ws = wb.active
        img = Image_openpyxl(image_bytearray)
        ws.add_image(img,'F39') 

        ws['A2'] = dados_bobina.loc['sap'] 
        ws['A3'] = dados_bobina.loc['descricao'] 
        ws['A5'] = dados_bobina.loc['conferente'] 
        ws['A9'] = dados_bobina.loc['lote'] 
        ws['D9'] = dados_bobina.loc['data'] 
        ws['A18'] = str(dados_bobina.loc['quantidade'])
        ws['A39'] = dados_bobina.loc['tipo'] #.replace('BOBINA ALUMINIO ', '')

    if dados_bobina.loc['tipo_de_etiqueta'] == 'BLOQUEADO':
        imagem_bobina_qr = qrcode.make(texto_qrcode , version=10, box_size=3, border=2, error_correction=qrcode.constants.ERROR_CORRECT_H) #, fit=True)
        image_bytearray = io.BytesIO()
        imagem_bobina_qr.save(image_bytearray, format='PNG', name='qrcode.png')        
        wb = load_workbook('BLOQUEADO.xlsx')
        ws = wb.active        
        img = Image_openpyxl(image_bytearray)
        ws.add_image(img,'F15')

        ws['C2'] = str(dados_bobina.loc['sap']) + ' - ' +  dados_bobina.loc['descricao'] #codigo do produto
        ws['A3'] = dados_bobina.loc['quantidade'] #quantidade do produto
        ws['A5'] = dados_bobina.loc['lote'] #lote do produto
        ws['A13'] = dados_bobina.loc['data'] #data de entrada do produto

    wb.save('Etiqueta_download.xlsx')
    stream = BytesIO()
    wb.save(stream)
    towrite = stream.getvalue()
    b64 = base64.b64encode(towrite).decode()  # some strings

    # link para download e nome do arquivo
    linko = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="etiqueta.xlsx">Download etiqueta</a>'
    st.markdown(linko, unsafe_allow_html=True)


def etiquetas_bobinas() -> None:
    st.subheader('Etiquetas de bobinas')

    doc_ref = db.collection('bobinas').document('bobinas')
    doc = doc_ref.get()

    if doc.exists:
        dicionario = doc.to_dict()
        csv = dicionario['dataframe']

        csv_string = StringIO(csv)
        df_bobinas = pd.read_csv(csv_string, sep=',')

        data_etiqueta = st.date_input('Data da etiqueta')
        status_etiqueta = st.selectbox('Tipo de etiqueta', ['Liberado', 'Não conforme'])

        df_bobinas['data'] = pd.to_datetime(df_bobinas['data']).dt.date
        df_etiqueta_dia = df_bobinas.loc[(df_bobinas['data'] == data_etiqueta) & (df_bobinas['status'] == status_etiqueta)]

        if df_etiqueta_dia.empty:
            st.warning('Não há bobinas para a data e tipo de etiqueta informado')
        else:
            lista_etiquetas = list(df_etiqueta_dia.index)

            for bobina in lista_etiquetas:
                texto_expander = ''.join(('Lote: ', str(df_etiqueta_dia.loc[bobina]['lote']), ' Quantidade: ', str(df_etiqueta_dia.loc[bobina]['quantidade']), ' (', str(bobina), ')'))
                with st.expander(texto_expander):
                    texto_qrcode = ''
                    for colunas in df_etiqueta_dia.columns:
                        if colunas != 'tipo_de_etiqueta':
                            texto_qrcode = ''.join((texto_qrcode, str(df_etiqueta_dia.loc[bobina, colunas]), ','))
                            st.write(f'**{colunas}:** {df_etiqueta_dia.loc[bobina, colunas]}')

                    texto_qrcode = texto_qrcode[0:-1]
                    download_etiqueta(texto_qrcode, df_bobinas.iloc[bobina])

    
def login_session_state() -> None:
    senha = st.secrets['pass']
    senha_input = st.text_input('Senha:', type='password')

    botao_logar = st.button('Logar')

    if botao_logar:
        if senha_input == senha:
            st.session_state['logado'] = True
            st.experimental_rerun()
        else:
            st.error('Senha incorreta')


if __name__ == "__main__":
    c1,c2 = st.sidebar.columns([1,1])
    c1.image('logo2.png', width=150)

    st.sidebar.subheader('Bobinas')
    telas_bobinas = ['Entrada de bobinas', 'Etiquetas', 'Inventário']
    tela_bobina = st.sidebar.radio('Menu bobinas', telas_bobinas)

    if tela_bobina == 'Entrada de bobinas':
        entrada_bobinas()

    if tela_bobina == 'Inventário':
        tela_inventario = st.sidebar.radio('Opções de inventário', ['Inserir', 'Visualizar']) #'Importar',

        if tela_inventario == 'Inserir':
            inserir_inventario()

        if tela_inventario == 'Visualizar':
            visualizar_inventario()

    if tela_bobina == 'Etiquetas':
        etiquetas_bobinas()
